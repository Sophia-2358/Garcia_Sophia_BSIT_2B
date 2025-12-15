import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from datetime import datetime
import os

class PlantScheduler:
    def __init__(self, root):
        self.root = root
        self.root.title("Plant Watering Scheduler")
        self.root.geometry("600x400")
        self.root.configure(bg="#2d3436") 
        self.file = "plants.xlsx"
        
        if not os.path.exists(self.file):
            wb = openpyxl.Workbook()
            wb.active.append(["Plant Name", "Days", "Last Watered", "Time"])
            wb.save(self.file)
        
        # Show home page muna
        self.show_home_page()
    
    def show_home_page(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        
        home_frame = tk.Frame(self.root, bg="#2d3436")
        home_frame.pack(fill=tk.BOTH, expand=True)
        
        # Welcome messages
        welcome_label = tk.Label(home_frame, text="🌱 Welcome User! 🌱",  #title
                                bg="#2d3436", fg="#dfe6e9", 
                                font=("Arial", 24, "bold"))
        welcome_label.pack(pady=80)
        
        subtitle = tk.Label(home_frame, text="Plant Watering Scheduler", #subtitle
                           bg="#2d3436", fg="#dfe6e9", 
                           font=("Arial", 14))
        subtitle.pack(pady=10)
        
        start_btn = tk.Button(home_frame, text="Get Started", 
                             command=self.show_main_page, # CALL FUNCTION TO SHOW MAIN PAGE
                             bg="#00b85c", fg="white", 
                             font=("Arial", 12, "bold"),
                             width=20, height=2,
                             activebackground="#00cc66")
        start_btn.pack(pady=30)
    
    def show_main_page(self):
        for widget in self.root.winfo_children():
            widget.destroy()
        
        self.create_widgets()
        self.load()
    
    def create_widgets(self):
        # Input section
        f = tk.Frame(self.root, bg="#141718", padx=10, pady=10) 
        f.pack(fill=tk.X)
        tk.Label(f, text="Plant:", bg="#141718", fg="#dfe6e9", font=("Arial", 10)).pack(side=tk.LEFT) 
        self.name = tk.Entry(f, width=15)  # plant name
        self.name.pack(side=tk.LEFT, padx=5)
        tk.Label(f, text="Days:", bg="#141718", fg="#dfe6e9", font=("Arial", 10)).pack(side=tk.LEFT) 
        self.days = tk.Entry(f, width=8)  # : days ng pagdilig ng halaman
        self.days.pack(side=tk.LEFT, padx=5)
        tk.Button(f, text="Add", command=self.add, bg="#00b85c", fg="white", font=("Arial", 9, "bold"), activebackground="#00cc66").pack(side=tk.LEFT) #green button
        
        # Table Section
        self.tree = ttk.Treeview(self.root, columns=("Plant", "Days", "Last", "Time", "Left"), show="headings", height=10) 
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col) 
        self.tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Buttons section
        bf = tk.Frame(self.root, bg="#141718", pady=10) 
        bf.pack(fill=tk.X)
        
        # Added inner frame para ma-center yung buttons
        btn_container = tk.Frame(bf, bg="#141718")  #Dark background for button container
        btn_container.pack(expand=True)  #expand=True para ma-center yung buttons
        
        #Buttons naka-pack sa btn_container instead of bf, with width=12 para same size
        tk.Button(btn_container, text="💧 Water", command=self.water, bg="#00b85c", fg="white", font=("Arial", 9, "bold"), width=12, activebackground="#00cc66").pack(side=tk.LEFT, padx=5)  # CHANGED: Green themed button
        tk.Button(btn_container, text="🗑️ Delete", command=self.delete, bg="#00b85c", fg="white", font=("Arial", 9, "bold"), width=12, activebackground="#00cc66").pack(side=tk.LEFT, padx=5)  # CHANGED: Green themed button
        tk.Button(btn_container, text="🔄 Refresh", command=self.load, bg="#00b85c", fg="white", font=("Arial", 9, "bold"), width=12, activebackground="#00cc66").pack(side=tk.LEFT, padx=5)  # CHANGED: Green themed button
        tk.Button(btn_container, text="🏠 Home", command=self.show_home_page, bg="#00b85c", fg="white", font=("Arial", 9, "bold"), width=12, activebackground="#00cc66").pack(side=tk.LEFT, padx=5)  # CHANGED: Home button to go back

    def add(self):
        n, d = self.name.get().strip(), self.days.get().strip()
        if not n or not d: return messagebox.showwarning("Error", "Fill all fields") 
        wb = openpyxl.load_workbook(self.file)  #Open Excel file
        now = datetime.now()  #Get ang current date
        wb.active.append([n, int(d), now.strftime("%Y-%m-%d"), now.strftime("%I:%M %p")]) 
        wb.save(self.file)  #Save Excel file
        self.name.delete(0, tk.END)  
        self.days.delete(0, tk.END) 
        self.load()  

    def load(self):
        for i in self.tree.get_children(): self.tree.delete(i) 
        wb = openpyxl.load_workbook(self.file)  # Open Excel file
        for row in wb.active.iter_rows(min_row=2, values_only=True): 
            if row[0]:  # Check if yung plant name ay meron
                left = row[1] - (datetime.now() - datetime.strptime(row[2], "%Y-%m-%d")).days  #Calculate days left
                time = row[3] if len(row) > 3 and row[3] else "N/A" 
                self.tree.insert("", tk.END, values=(row[0], f"{row[1]}d", row[2], time, f"{left}d")) 

    def water(self):
        if not (s := self.tree.selection()): return 
        wb = openpyxl.load_workbook(self.file)  
        plant = self.tree.item(s[0])['values'][0]  
        now = datetime.now()  
        for row in wb.active.iter_rows(min_row=2):  
            if row[0].value == plant:  
                row[2].value = now.strftime("%Y-%m-%d") 
                row[3].value = now.strftime("%I:%M %p")  
                break
        wb.save(self.file)  #Save Excel file
        self.load()  #Refresh table display
        messagebox.showinfo("Success", f"✅ {plant} watered!")

    def delete(self):
        if not (s := self.tree.selection()): return  
        wb = openpyxl.load_workbook(self.file) 
        plant = self.tree.item(s[0])['values'][0]  
        for i, row in enumerate(wb.active.iter_rows(min_row=2), start=2): 
            if row[0].value == plant:  
                wb.active.delete_rows(i)  # Delete row from Excel
                break
        wb.save(self.file)  #Save Excel file
        self.load()  #Refresh ng mga table

root = tk.Tk()
PlantScheduler(root)
root.mainloop()